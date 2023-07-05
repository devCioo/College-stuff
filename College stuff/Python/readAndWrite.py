from datetime import datetime
import random
from tempfile import mktemp
from os import unlink
from pickle import dump, load
import pandas as pd
from fastparquet import ParquetFile
from openpyxl import Workbook, load_workbook
from itertools import product

def saveAndLoad(func):
    def wrapper(*args):
        print(f"Czas zapisu i odczytu {args[0]} elementów modułem {func.__name__}:")
        return func(*args)
    return wrapper

@saveAndLoad
def pickle(elems: int):
    temp_file = mktemp()
    collection = []
    for i in range(elems):
        collection.append(random.randint(1,100))
    with open(temp_file, 'wb') as f:
        t1 = datetime.now()
        dump(collection, f)
        t2 = datetime.now()
        svtime = (t2 - t1).total_seconds()
        f.flush()
    with open(temp_file, 'rb') as f:
        t1 = datetime.now()
        extr_file = load(f)
        t2 = datetime.now()
        ldtime = (t2 - t1).total_seconds()
    unlink(temp_file)
    return f"[{svtime}s, {ldtime}s]"

@saveAndLoad
def parquet(elems: int):
    temp_file = mktemp()
    dictionary = {}
    for i in range(elems):
        dictionary[i]: random.randint(1,100)
    df = pd.DataFrame.from_dict(dictionary)
    t1 = datetime.now()
    df.to_parquet(temp_file, compression='GZIP')
    pf = ParquetFile(temp_file)
    t2 = datetime.now()
    svtime = (t2 - t1).total_seconds()
    t1 = datetime.now()
    df = pf.to_pandas()
    t2 = datetime.now()
    ldtime = (t2 - t1).total_seconds()
    unlink(temp_file)
    return f"[{svtime}s, {ldtime}s]"

@saveAndLoad
def xlsx(elems: int):
    file_name = mktemp(suffix='.xlsx')
    data = {}
    for i in range(elems):
        data[f'A{i}']: random.randint(1,100)
    t1 = datetime.now()
    wb = Workbook()
    sheet = wb.create_sheet('test')
    for cell_key, cell_value in data.items():
        sheet[cell_key] = cell_value
    wb.save(file_name)
    t2 = datetime.now()
    svtime = (t2 - t1).total_seconds()
    t1 = datetime.now()
    wb = load_workbook(file_name)
    t2 = datetime.now()
    ldtime = (t2 - t1).total_seconds()
    unlink(file_name)
    return f"[{svtime}s, {ldtime}s]"

print(pickle(100)), print(parquet(100)), print(xlsx(100), "\n")
print(pickle(10000)), print(parquet(10000)), print(xlsx(10000), "\n")
print(pickle(100000)), print(parquet(100000)), print(xlsx(100000))